#!/usr/bin/env python3
"""
Network Configuration Comparison Tool
Compares new_output.json with whitelist_file.json and generates Excel report
"""

import json
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import traceback
import sys

def load_json_file(file_path):
    """Safely load JSON file with error handling"""
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        print(f"✓ Successfully loaded: {file_path}")
        return data
    except json.JSONDecodeError as e:
        print(f"✗ JSON decode error in {file_path}: {e}")
        return None
    except Exception as e:
        print(f"✗ Error loading {file_path}: {e}")
        return None

def normalize_text(text):
    """Trim whitespace and normalize text for comparison"""
    if isinstance(text, str):
        return text.strip()
    return text

def compare_configurations(output_data, whitelist_data, hostname):
    """
    Compare configuration sections and identify missing/additional configs
    Returns lists of missing and additional configurations
    """
    missing_configs = []
    additional_configs = []

    try:
        # Get sections from both files
        output_sections = output_data.get(hostname, {})

        # Compare each section in whitelist with output
        for section_name, whitelist_section in whitelist_data.items():
            try:
                # Get the corresponding section from output
                output_section = output_sections.get(section_name, [])

                # Handle different data types
                if isinstance(whitelist_section, dict) and 'must_include' in whitelist_section:
                    required_items = whitelist_section['must_include']

                    # Convert output section to normalized strings for comparison
                    if isinstance(output_section, list):
                        output_items = [normalize_text(str(item)) for item in output_section if item]
                    else:
                        output_items = [normalize_text(str(output_section))] if output_section else []

                    # Check for missing required configurations
                    for required_item in required_items:
                        required_normalized = normalize_text(str(required_item))
                        if required_normalized and required_normalized not in output_items:
                            missing_configs.append(f"missing config: {required_item}:{hostname}")

                    # Check for additional configurations (items in output but not in whitelist)
                    for output_item in output_items:
                        output_normalized = normalize_text(str(output_item))
                        if output_normalized and output_normalized not in [normalize_text(str(item)) for item in required_items]:
                            # Only flag non-empty, meaningful additional configs
                            if len(output_normalized) > 0 and not output_normalized.startswith('!') and not output_normalized.startswith('#'):
                                additional_configs.append(f"additional config: {output_item}:{hostname}")

            except Exception as section_error:
                print(f"Warning: Error processing section '{section_name}': {section_error}")
                continue

    except Exception as e:
        print(f"Error during comparison for {hostname}: {e}")

    return missing_configs, additional_configs

def create_excel_report(output_data, whitelist_data, output_path="config_comparison.xlsx"):
    """Create Excel report with comparison results"""
    try:
        # Create workbook with multiple sheets
        wb = Workbook()

        # Remove default sheet and create our sheets
        wb.remove(wb.active)
        sheet1 = wb.create_sheet("Device Summary", 0)
        sheet2 = wb.create_sheet("Comparison Results", 1)

        # Sheet 1: Device Summary
        sheet1_data = []
        sheet2_data = []
        sr_no = 1

        # Process each device in output data
        for hostname, device_config in output_data.items():
            try:
                # Get configuration sections for this device
                sections = list(device_config.keys())

                # Add rows for each section to sheet 1
                for section in sections:
                    sheet1_data.append({
                        'SR.No.': sr_no,
                        'hostname': hostname,
                        'section': section
                    })
                    sr_no += 1

                # Compare configurations and get results
                missing_configs, additional_configs = compare_configurations(
                    output_data, whitelist_data, hostname
                )

                # Add results to sheet 2
                for config in missing_configs:
                    sheet2_data.append({'to_check': config})

                for config in additional_configs:
                    sheet2_data.append({'to_check': config})

            except Exception as device_error:
                print(f"Warning: Error processing device '{hostname}': {device_error}")
                # Add error entry to sheet 2
                sheet2_data.append({'to_check': f"error processing device: {hostname}"})
                continue

        # Create DataFrames and write to Excel
        if sheet1_data:
            df1 = pd.DataFrame(sheet1_data)
            for r in dataframe_to_rows(df1, index=False, header=True):
                sheet1.append(r)
        else:
            # Add headers even if no data
            sheet1.append(['SR.No.', 'hostname', 'section'])

        if sheet2_data:
            df2 = pd.DataFrame(sheet2_data)
            for r in dataframe_to_rows(df2, index=False, header=True):
                sheet2.append(r)
        else:
            # Add headers even if no data
            sheet2.append(['to_check'])

        # Format headers
        for sheet in [sheet1, sheet2]:
            for cell in sheet[1]:  # First row is header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
                cell.alignment = Alignment(horizontal='center')

        # Auto-adjust column widths
        for sheet in [sheet1, sheet2]:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Max width of 50
                sheet.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(output_path)
        print(f"✓ Excel report saved as: {output_path}")
        print(f"  - Sheet 1: {len(sheet1_data)} device section entries")
        print(f"  - Sheet 2: {len(sheet2_data)} comparison results")

    except Exception as e:
        print(f"✗ Error creating Excel report: {e}")
        traceback.print_exc()

def main():
    """Main function to orchestrate the comparison process"""
    print("Network Configuration Comparison Tool")
    print("="*50)

    # File paths - modify these as needed
    output_file_path = "//home//suneelk//ansible_outputs//remediation//au-br2-adm-corp-csw01_output.json"
    whitelist_file_path = "//home//suneelk//ansible_outputs//white_list.json"
    excel_output_path = "//home//suneelk//ansible_outputs//config_comparison.xlsx"

    try:
        # Load JSON files
        print("\nStep 1: Loading configuration files...")
        output_data = load_json_file(output_file_path)
        whitelist_data = load_json_file(whitelist_file_path)

        if not output_data:
            print(f"✗ Failed to load output file: {output_file_path}")
            return

        if not whitelist_data:
            print(f"✗ Failed to load whitelist file: {whitelist_file_path}")
            return

        print(f"\nStep 2: Analyzing configurations...")
        print(f"  - Output file devices: {len(output_data)}")
        print(f"  - Whitelist sections: {len(whitelist_data)}")

        # Create Excel report
        print(f"\nStep 3: Generating Excel report...")
        create_excel_report(output_data, whitelist_data, excel_output_path)

        print(f"\n✓ Configuration comparison completed successfully!")
        print(f"Please check the Excel file: {excel_output_path}")

    except Exception as e:
        print(f"\n✗ Error in main process: {e}")
        traceback.print_exc()
        return 1

    return 0

if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)
